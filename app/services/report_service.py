"""Report Service — generates PDF, Excel, and CSV reports."""
import csv
import io
import os
from datetime import datetime, date, timedelta
from app.extensions import db
from app.models.report import Report
from app.services.attendance_service import AttendanceService


class ReportService:
    """Generates and manages attendance reports."""

    @staticmethod
    def generate_csv_report(period='daily'):
        """Generate a CSV attendance report."""
        from app.models.attendance import Attendance
        from app.models.student import Student

        today = date.today()
        if period == 'daily':
            start_date = today
        elif period == 'weekly':
            start_date = today - timedelta(days=today.weekday())
        elif period == 'monthly':
            start_date = today.replace(day=1)
        else:
            start_date = today - timedelta(days=180)

        records = db.session.query(
            Student.name, Student.roll_number, Student.department,
            Attendance.date, Attendance.status, Attendance.time_in
        ).join(Student, Attendance.student_id == Student.id).filter(
            Attendance.date >= start_date
        ).order_by(Attendance.date, Student.roll_number).all()

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Name', 'Roll Number', 'Department', 'Date', 'Status', 'Time In'])

        for r in records:
            writer.writerow([
                r.name, r.roll_number, r.department,
                r.date.strftime('%Y-%m-%d'), r.status,
                r.time_in.strftime('%H:%M') if r.time_in else 'N/A'
            ])

        output.seek(0)
        return output.getvalue()

    @staticmethod
    def generate_excel_report(period='daily'):
        """Generate an Excel attendance report."""
        from app.models.attendance import Attendance
        from app.models.student import Student

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return None

        today = date.today()
        if period == 'daily':
            start_date = today
        elif period == 'weekly':
            start_date = today - timedelta(days=today.weekday())
        elif period == 'monthly':
            start_date = today.replace(day=1)
        else:
            start_date = today - timedelta(days=180)

        wb = Workbook()
        ws = wb.active
        ws.title = f'{period.capitalize()} Attendance Report'

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = f'AI Classroom Intelligence — {period.capitalize()} Attendance Report'
        ws['A1'].font = Font(name='Arial', size=16, bold=True, color='1a237e')
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:F2')
        ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")} | Period: {start_date} to {today}'
        ws['A2'].font = Font(name='Arial', size=10, italic=True)
        ws['A2'].alignment = Alignment(horizontal='center')

        # Headers
        headers = ['Name', 'Roll Number', 'Department', 'Date', 'Status', 'Time In']
        header_fill = PatternFill(start_color='1a237e', end_color='1a237e', fill_type='solid')
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Data
        records = db.session.query(
            Student.name, Student.roll_number, Student.department,
            Attendance.date, Attendance.status, Attendance.time_in
        ).join(Student, Attendance.student_id == Student.id).filter(
            Attendance.date >= start_date
        ).order_by(Attendance.date, Student.roll_number).all()

        status_colors = {
            'present': PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid'),
            'absent': PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid'),
            'late': PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid'),
        }

        for i, r in enumerate(records, 5):
            ws.cell(row=i, column=1, value=r.name)
            ws.cell(row=i, column=2, value=r.roll_number)
            ws.cell(row=i, column=3, value=r.department)
            ws.cell(row=i, column=4, value=r.date.strftime('%Y-%m-%d'))
            status_cell = ws.cell(row=i, column=5, value=r.status.capitalize())
            status_cell.fill = status_colors.get(r.status, PatternFill())
            ws.cell(row=i, column=6, value=r.time_in.strftime('%H:%M') if r.time_in else 'N/A')

        # Auto-width columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)

        # Summary sheet
        ws2 = wb.create_sheet('Summary')
        stats = AttendanceService.get_today_stats()
        ws2['A1'] = 'Summary Statistics'
        ws2['A1'].font = Font(size=14, bold=True)
        summary_data = [
            ('Total Students', stats['total_students']),
            ('Present Today', stats['present_today']),
            ('Absent Today', stats['absent_today']),
            ('Attendance Rate', f"{stats['attendance_rate']}%"),
        ]
        for i, (label, value) in enumerate(summary_data, 3):
            ws2.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws2.cell(row=i, column=2, value=value)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def generate_pdf_report(period='daily'):
        """Generate a PDF attendance report."""
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.units import inch
        except ImportError:
            return None

        from app.models.attendance import Attendance
        from app.models.student import Student

        today = date.today()
        if period == 'daily':
            start_date = today
        elif period == 'weekly':
            start_date = today - timedelta(days=today.weekday())
        elif period == 'monthly':
            start_date = today.replace(day=1)
        else:
            start_date = today - timedelta(days=180)

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle', parent=styles['Title'],
            fontSize=20, textColor=colors.HexColor('#1a237e'), spaceAfter=12
        )
        elements.append(Paragraph('AI Classroom Intelligence System', title_style))
        elements.append(Paragraph(
            f'{period.capitalize()} Attendance Report — {start_date} to {today}',
            styles['Normal']
        ))
        elements.append(Spacer(1, 20))

        # Statistics
        stats = AttendanceService.get_today_stats()
        stats_data = [
            ['Total Students', 'Present', 'Absent', 'Attendance Rate'],
            [str(stats['total_students']), str(stats['present_today']),
             str(stats['absent_today']), f"{stats['attendance_rate']}%"]
        ]
        stats_table = Table(stats_data, colWidths=[2 * inch] * 4)
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f5f5f5')),
        ]))
        elements.append(stats_table)
        elements.append(Spacer(1, 20))

        # Attendance table
        records = db.session.query(
            Student.name, Student.roll_number, Student.department,
            Attendance.date, Attendance.status
        ).join(Student, Attendance.student_id == Student.id).filter(
            Attendance.date >= start_date
        ).order_by(Attendance.date, Student.roll_number).limit(200).all()

        table_data = [['Name', 'Roll Number', 'Department', 'Date', 'Status']]
        for r in records:
            table_data.append([
                r.name, r.roll_number, r.department,
                r.date.strftime('%Y-%m-%d'), r.status.capitalize()
            ])

        if len(table_data) > 1:
            t = Table(table_data, colWidths=[2.5 * inch, 1.5 * inch, 1.5 * inch, 1.5 * inch, 1 * inch])
            style_commands = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
            ]
            t.setStyle(TableStyle(style_commands))
            elements.append(t)

        # Footer
        elements.append(Spacer(1, 30))
        elements.append(Paragraph(
            f'Report generated on {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} | AI Classroom Intelligence System',
            styles['Normal']
        ))

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()
